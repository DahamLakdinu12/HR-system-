#!/usr/bin/env python3
"""Convert the approved HR workbook into a SQL Server bulk-import file."""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "Paycode",
    "Sex",
    "FirstName",
    "LastName",
    "DateOfBirth",
    "DateJoined",
    "DateOfPromotion",
    "IncrementLevel",
    "SalaryPoint",
    "StagnationAllowance",
    "DateOfRetired",
    "NextIncrementDate",
    "NearestPolice",
    "PostDescription",
    "EPFNumber",
    "SalaryScale",
    "StartPoint",
    "Increment",
    "Seq",
    "NoOfIncrement",
    "E.GradeCode",
    "NewGrade",
    "W.AbbrName",
    "DeptDescription",
]

PAYABLE_SALARY_HEADERS = {"NearestPolice", "NearestPolice (Payable 2026)"}
BASIC_SALARY_2027_HEADERS = {"EPFNumber", "EPFNumber (2027 Basic)"}

DATE_COLUMNS = {
    "DateOfBirth",
    "DateJoined",
    "DateOfPromotion",
    "DateOfRetired",
    "NextIncrementDate",
}

DECIMAL_COLUMNS = {
    "SalaryPoint",
    "StagnationAllowance",
    "StartPoint",
    "Increment",
}

INTEGER_COLUMNS = {"IncrementLevel", "Seq", "NoOfIncrement"}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()


def clean_date(value: object, column: str, row_number: int) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = clean_text(value)
    for pattern in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Row {row_number}: invalid {column} value {value!r}")


def clean_decimal(value: object, column: str, row_number: int) -> str:
    if value in (None, ""):
        return ""
    try:
        return format(Decimal(clean_text(value).replace(",", "")), "f")
    except InvalidOperation as error:
        raise ValueError(f"Row {row_number}: invalid {column} value {value!r}") from error


def clean_allowance(value: object) -> tuple[str, str]:
    text = clean_text(value)
    if not text:
        return "", ""
    try:
        return format(Decimal(text.replace(",", "")), "f"), ""
    except InvalidOperation:
        return "", text


def clean_payable_salary(value: object) -> str:
    """Extract payable 2026 salary from prefixed or numeric column M values."""
    text = clean_text(value)
    if text.upper().startswith("PS"):
        text = text[2:].strip()

    try:
        return format(Decimal(text.replace(",", "")), "f")
    except InvalidOperation:
        return ""


def clean_basic_salary_2027(value: object) -> str:
    """Extract 2027 basic salary from values such as 'BS 95770'."""
    text = clean_text(value)
    if text.upper().startswith("BS"):
        text = text[2:].strip()

    try:
        return format(Decimal(text.replace(",", "")), "f")
    except InvalidOperation:
        return ""


def clean_integer(value: object, column: str, row_number: int) -> str:
    if value in (None, ""):
        return ""
    try:
        return str(int(Decimal(clean_text(value))))
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"Row {row_number}: invalid {column} value {value!r}") from error


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    worksheet = load_workbook(args.workbook, data_only=True, read_only=True).active
    headers = [clean_text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    normalized_headers = headers.copy()
    if len(normalized_headers) > 12 and normalized_headers[12] in PAYABLE_SALARY_HEADERS:
        normalized_headers[12] = "NearestPolice"
    if len(normalized_headers) > 14 and normalized_headers[14] in BASIC_SALARY_2027_HEADERS:
        normalized_headers[14] = "EPFNumber"
    if normalized_headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected workbook columns: {headers!r}")

    seen_pay_codes: set[str] = set()
    imported_rows = 0
    payable_salary_fallbacks = 0
    missing_basic_salary_2027 = 0
    args.output.parent.mkdir(parents=True, exist_ok=True)

    with args.output.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(
            output,
            delimiter="\t",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )

        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in cells]
            if not any(value not in (None, "") for value in values):
                continue

            record = dict(zip(normalized_headers, values))
            pay_code = clean_text(record["Paycode"])
            if not pay_code:
                raise ValueError(f"Row {row_number}: Paycode is required")
            if pay_code in seen_pay_codes:
                raise ValueError(f"Row {row_number}: duplicate Paycode {pay_code}")
            seen_pay_codes.add(pay_code)

            normalized: list[str] = []
            for column in EXPECTED_HEADERS:
                value = record[column]
                if column == "StagnationAllowance":
                    allowance, allowance_note = clean_allowance(value)
                    normalized.extend((allowance, allowance_note))
                elif column == "NearestPolice":
                    payable_salary = clean_payable_salary(value)
                    if not payable_salary:
                        payable_salary = clean_decimal(
                            record["SalaryPoint"], "SalaryPoint", row_number
                        )
                        payable_salary_fallbacks += 1
                    normalized.append(payable_salary)
                elif column == "EPFNumber":
                    basic_salary_2027 = clean_basic_salary_2027(value)
                    if not basic_salary_2027:
                        missing_basic_salary_2027 += 1
                    normalized.append(basic_salary_2027)
                elif column in DATE_COLUMNS:
                    normalized.append(clean_date(value, column, row_number))
                elif column in DECIMAL_COLUMNS:
                    normalized.append(clean_decimal(value, column, row_number))
                elif column in INTEGER_COLUMNS:
                    normalized.append(clean_integer(value, column, row_number))
                else:
                    normalized.append(clean_text(value))

            writer.writerow(normalized)
            imported_rows += 1

    print(f"Imported {imported_rows} rows with {len(seen_pay_codes)} unique pay codes.")
    if payable_salary_fallbacks:
        print(
            f"Used SalaryPoint as the current-salary fallback for "
            f"{payable_salary_fallbacks} rows without a valid column M payable amount."
        )
    if missing_basic_salary_2027:
        print(
            f"Left {missing_basic_salary_2027} rows unmatched because column O "
            "does not contain a valid 2027 basic salary."
        )


if __name__ == "__main__":
    main()

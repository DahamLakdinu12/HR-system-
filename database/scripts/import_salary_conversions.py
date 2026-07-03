#!/usr/bin/env python3
"""Convert the approved consolidated salary workbook into a SQL import file."""

from __future__ import annotations

import argparse
import csv
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


WORKSHEET_NAME = "SalTables"
EXPECTED_HEADERS = (
    "SALARY STEP",
    "NEW BASIC SAL 2027",
    "NEWINCREVALUE",
    "BASIC SALARY APRIL 2025 (PAID AMOUNT 01.01.2025)",
    "BASIC SALARY 2026",
)

# The consolidated workbook repeats a scale title and, where applicable, a
# grade title before each table. Map those approved labels explicitly to the
# employee grade codes used by the HR staff workbook.
SECTION_GRADES = {
    ("HM 2-3 - 2025", ""): "HM-2-3",
    ("HM 2-1 - 2025", ""): "HM-2-1",
    ("HM 1-3 - 2025", ""): "HM-1-3",
    ("HM 1-1 - 2025", ""): "HM-1-1",
    ("MM 1-1 - 2025", "GRADE I (DD)"): "MM-1-1-I",
    ("MM 1-1 - 2025", "GRADE II (AD)"): "MM-1-1-II",
    ("JM 1-1 - 2025", "GRADE I"): "JM-1-1-I",
    ("JM 1-1 - 2025", "GRADE II"): "JM-1-1-II",
    ("MA 2-2 - 2025", "GRADE I"): "MA-2-2-I",
    ("MA 2-2 - 2025", "GRADE II"): "MA-2-2-II",
    ("MA 2-2 - 2025", "GRADE III"): "MA-2-2-III",
    ("MA 1-2 - 2025", "GRADE I"): "MA-1-2-I",
    ("MA 1-2 - 2025", "GRADE II"): "MA-1-2-II",
    ("MA 1-2 - 2025", "GRADE III"): "MA-1-2-III",
    ("MA 1-1 - 2025", "GRADE II"): "MA-1-1",
    ("PL 3 - 2025", "GRADE I"): "PL-3-I",
    ("PL 3 - 2025", "GRADE II"): "PL-3-II",
    ("PL 3 - 2025", "GRADE III"): "PL-3-III",
    ("PL 2 - 2025", "GRADE I"): "PL-2-I",
    ("PL 2 - 2025", "GRADE II"): "PL-2-II",
    ("PL 2 - 2025", "GRADE III"): "PL-2-III",
    ("PL 1 - 2025", "GRADE I"): "PL-1-I",
    ("PL 1 - 2025", "GRADE II"): "PL-1-II",
    ("PL 1 - 2025", "GRADE III"): "PL-1-III",
}

# Starting salary and increment segments transcribed from the approved
# Salary Scale.pdf. A segment is (number of annual increments, increment value).
# The starting salary is the first point, so 18 increments produce 19 points.
SALARY_SCALE_RULES = {
    "HM-2-3": (173000, ((12, 4850),)),
    "HM-2-1": (161140, ((12, 4850),)),
    "HM-1-3": (152500, ((15, 4100),)),
    "HM-1-1": (140640, ((15, 4100),)),
    "MM-1-1-I": (119940, ((14, 3450),)),
    "MM-1-1-II": (91690, ((16, 2480),)),
    "JM-1-1-I": (88290, ((17, 2040),)),
    "JM-1-1-II": (72650, ((15, 1360),)),
    "MA-1-2-I": (61430, ((7, 1080), (12, 1280))),
    "MA-1-2-II": (52250, ((6, 630), (12, 1080))),
    "MA-1-2-III": (46220, ((18, 540),)),
    "MA-2-2-I": (65950, ((19, 1280),)),
    "MA-2-2-II": (56570, ((6, 630), (14, 1080))),
    "MA-2-2-III": (50540, ((18, 540),)),
    "MA-1-1": (52250, ((6, 630), (12, 890))),
    "PL-3-I": (54170, ((9, 590), (12, 630))),
    "PL-3-II": (48720, ((14, 540),)),
    "PL-3-III": (43280, ((18, 490),)),
    "PL-2-I": (53190, ((9, 590), (12, 630))),
    "PL-2-II": (47740, ((14, 540),)),
    "PL-2-III": (42300, ((18, 490),)),
    "PL-1-I": (50440, ((6, 540), (15, 590))),
    "PL-1-II": (45490, ((14, 490),)),
    "PL-1-III": (40500, ((18, 450),)),
}


def normalize_label(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().upper().split())


def decimal_value(value: object, section: str, row: int, column: str) -> Decimal:
    if value is None:
        raise ValueError(f"{section} row {row}: {column} is required")
    try:
        return Decimal(str(value).replace(",", ""))
    except InvalidOperation as error:
        raise ValueError(
            f"{section} row {row}: invalid {column} value {value!r}"
        ) from error


def output_decimal(value: Decimal) -> str:
    return format(value, "f")


def salary_scale_points(grade_code: str) -> list[Decimal]:
    start, segments = SALARY_SCALE_RULES[grade_code]
    points = [Decimal(start)]
    for increment_count, increment_value in segments:
        for _ in range(increment_count):
            points.append(points[-1] + Decimal(increment_value))
    return points


def read_sections(workbook_path: Path) -> tuple[list[list[str]], int]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if WORKSHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Missing worksheet: {WORKSHEET_NAME}")

    worksheet = workbook[WORKSHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    current_scale = ""
    current_grade = ""
    imported_sections: set[str] = set()
    output_rows: list[list[str]] = []
    stagnation_extension_rows = 0

    for index, row in enumerate(rows):
        label = normalize_label(row[0])
        if not label:
            continue
        if label.startswith("GRADE "):
            current_grade = label
            continue
        if label.endswith("- 2025"):
            current_scale = label
            current_grade = ""
            continue
        if label != "SALARY STEP":
            continue

        headers = tuple(normalize_label(value) for value in row[:5])
        if headers != EXPECTED_HEADERS:
            raise ValueError(
                f"Unexpected salary headers at row {index + 1}: {headers!r}"
            )

        section_key = (current_scale, current_grade)
        grade_code = SECTION_GRADES.get(section_key)
        if grade_code is None:
            raise ValueError(f"Unmapped salary section: {section_key!r}")
        if grade_code in imported_sections:
            raise ValueError(f"Duplicate salary section for {grade_code}")
        imported_sections.add(grade_code)

        section_rows: list[tuple[int, tuple[object, ...]]] = []
        data_index = index + 1
        while data_index < len(rows) and rows[data_index][0] is not None:
            section_rows.append((data_index + 1, rows[data_index]))
            data_index += 1
        if not section_rows:
            raise ValueError(f"{grade_code}: salary section has no points")

        approved_points = salary_scale_points(grade_code)
        normal_point_count = len(approved_points)
        if len(section_rows) < normal_point_count:
            raise ValueError(
                f"{grade_code}: expected at least {normal_point_count} points "
                f"from Salary Scale.pdf, found {len(section_rows)}"
            )
        stagnation_extension_rows += len(section_rows) - normal_point_count
        final_increment = Decimal(SALARY_SCALE_RULES[grade_code][1][-1][1])

        previous_step: int | None = None
        for point_index, (row_number, point) in enumerate(section_rows):
            try:
                salary_step = int(Decimal(str(point[0])))
            except (InvalidOperation, ValueError) as error:
                raise ValueError(
                    f"{grade_code} row {row_number}: invalid salary step {point[0]!r}"
                ) from error
            if previous_step is not None and salary_step != previous_step + 1:
                raise ValueError(
                    f"{grade_code} row {row_number}: expected step "
                    f"{previous_step + 1}, found {salary_step}"
                )
            previous_step = salary_step

            basic_salary_2027 = decimal_value(
                point[1], grade_code, row_number, "New Basic sal 2027"
            )
            if point_index < normal_point_count:
                expected_salary = approved_points[point_index]
            else:
                expected_salary = approved_points[-1] + (
                    final_increment * (point_index - normal_point_count + 1)
                )
            if basic_salary_2027 != expected_salary:
                raise ValueError(
                    f"{grade_code} row {row_number}: salary progression requires "
                    f"{expected_salary}, found {basic_salary_2027}"
                )
            paid_salary_2025 = decimal_value(
                point[3], grade_code, row_number, "2025 paid salary"
            )
            basic_salary_2026 = decimal_value(
                point[4], grade_code, row_number, "2026 basic salary"
            )

            if point[2] is not None:
                increment_amount = decimal_value(
                    point[2], grade_code, row_number, "increment value"
                )
            elif point_index + 1 < len(section_rows):
                next_basic = decimal_value(
                    section_rows[point_index + 1][1][1],
                    grade_code,
                    section_rows[point_index + 1][0],
                    "New Basic sal 2027",
                )
                increment_amount = next_basic - basic_salary_2027
            else:
                increment_amount = final_increment

            output_rows.append([
                grade_code,
                f"{current_scale} {current_grade}".strip(),
                str(salary_step),
                output_decimal(basic_salary_2027),
                output_decimal(increment_amount),
                output_decimal(paid_salary_2025),
                output_decimal(basic_salary_2026),
                "1" if point_index >= normal_point_count else "0",
            ])

    missing_sections = set(SECTION_GRADES.values()) - imported_sections
    if missing_sections:
        raise ValueError(f"Missing salary sections: {sorted(missing_sections)}")
    return output_rows, stagnation_extension_rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    output_rows, stagnation_extension_rows = read_sections(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(
            output,
            delimiter="\t",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writerows(output_rows)

    print(
        f"Imported {len(output_rows)} salary points assigned to "
        f"{len(SECTION_GRADES)} employee grades."
    )
    if stagnation_extension_rows:
        print(
            f"Imported {stagnation_extension_rows} post-maximum rows as "
            "employee-specific stagnation extension points."
        )


if __name__ == "__main__":
    main()
